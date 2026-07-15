import pandas as pd
import openpyxl
import io
import os

def parse_imfu_file(file_content, filename):
    """
    Parse an IMFU excel file content bytes.
    Detects the header row dynamically, maps columns to a standard format,
    and returns a list of dictionaries representing the items.
    """
    ext = os.path.splitext(filename)[1].lower()
    df = None
    
    if ext == '.csv':
        try:
            df = pd.read_csv(io.BytesIO(file_content))
        except Exception as e:
            try:
                df = pd.read_csv(io.BytesIO(file_content), sep=';')
            except Exception as e2:
                raise ValueError(f"Impossible de lire le fichier CSV: {str(e2)}")
    else:
        # Try dynamic header with openpyxl (only works for xlsx)
        if ext in ['.xlsx', '.xlsm']:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                if not wb.sheetnames:
                    raise ValueError("Le fichier Excel ne contient aucune feuille.")
                ws = wb[wb.sheetnames[0]]
                
                max_f = 0
                header_row = 1
                for r in range(1, min(25, ws.max_row+1)):
                    f = sum(1 for c in range(1, ws.max_column+1) if ws.cell(r, c).value is not None)
                    if f > max_f:
                        max_f = f
                        header_row = r
                        
                df = pd.read_excel(io.BytesIO(file_content), sheet_name=0, header=header_row-1)
            except Exception as e:
                # Fallback
                df = None
        
        if df is None:
            # Fallback for .xls or if openpyxl failed
            try:
                # Try with default engine
                df = pd.read_excel(io.BytesIO(file_content), sheet_name=0)
            except Exception as e:
                try:
                    # Try with calamine or xlrd if .xls
                    engine = 'calamine' if ext == '.xls' else None
                    df = pd.read_excel(io.BytesIO(file_content), sheet_name=0, engine=engine)
                except Exception as e2:
                    raise ValueError(f"Format Excel non supporté ou fichier corrompu: {str(e2)}")

    if df is None or df.empty:
        raise ValueError("Le fichier est vide ou ne contient aucune donnée.")

    cols = df.columns
    def get_col(candidates):
        # Pass 1: Exact match (case-insensitive)
        for cand in candidates:
            for c in cols:
                c_str = str(c).lower().replace('\n', ' ').strip()
                if str(cand).lower().strip() == c_str:
                    return c
        # Pass 2: Substring match (case-insensitive)
        for cand in candidates:
            for c in cols:
                c_str = str(c).lower().replace('\n', ' ')
                if str(cand).lower() in c_str:
                    return c
        return None
        
    col_item = get_col(["item / sequence", "sequence", "item #", "n° ph", "id", "item", "n°"])
    col_scope = get_col(["scope", "périmètre", "perimetre", "system", "système"])
    col_status = get_col(["overall status", "status", "statut", "état", "etat"])
    col_priority = get_col(["complexity / priority", "priority", "complexity", "high/medium/low", "priorité", "priorite"])
    col_name = get_col(["name / description", "name", "description", "titre", "task", "tâche", "tache", "sujet", "subject"])
    
    col_owner = get_col(["responsible / pilote", "responsible", "owner", "responsable", "function", "métier", "pilote", "assignee"])
    col_action = get_col(["next action", "action", "prochaine etape", "next step"])
    col_issue = get_col(["blocker / issue", "blocker", "issue", "comment", "problème", "remarque", "point bloquant"])
    col_date = get_col(["planned end date", "planned date", "target date", "need date", "date cible", "échéance", "due date", "deadline"])
    
    col_progress = get_col(["overall progress", "progress", "avancement", "%"])
    
    items = []
    
    for idx, row in df.iterrows():
        name_val = str(row.get(col_name)).strip() if col_name and not pd.isna(row.get(col_name)) else ""
        
        owner_val = str(row.get(col_owner)).strip() if col_owner and not pd.isna(row.get(col_owner)) else ""
        if owner_val.lower() in ['nan', 'none', '']:
            owner_val = "Unassigned"
        else:
            owner_val = owner_val.title()
            
        # We don't skip entirely if name is missing if it's a fallback parse (header might be wrong)
        # But if name is missing AND everything else is missing, skip.
        if not name_val or name_val.lower() in ['nan', 'none', '']:
            # check if other important fields are there
            if owner_val == "Unassigned":
                continue
            name_val = "Sans titre"
            
        item_id = str(row.get(col_item)).strip() if col_item and not pd.isna(row.get(col_item)) else f"Row-{idx+2}"
        if item_id.lower() in ['nan', 'none', '']:
            item_id = f"Row-{idx+2}"
            
        item = {
            "source_file": filename,
            "item_id": item_id,
            "scope": str(row.get(col_scope)) if col_scope and not pd.isna(row.get(col_scope)) else "Unknown",
            "name": name_val,
            "status": "Not Started",
            "priority": "Medium",
            "owner": owner_val,
            "next_action": str(row.get(col_action)) if col_action and not pd.isna(row.get(col_action)) else "-",
            "issue": str(row.get(col_issue)) if col_issue and not pd.isna(row.get(col_issue)) else "-",
            "target_date": ""
        }
        
        if col_date and not pd.isna(row.get(col_date)):
            val = row.get(col_date)
            if isinstance(val, pd.Timestamp):
                item["target_date"] = val.strftime('%Y-%m-%d')
            else:
                item["target_date"] = str(val)[:10]
        
        if col_status and not pd.isna(row.get(col_status)) and str(row.get(col_status)).strip() != "":
            s = str(row.get(col_status)).strip().lower()
            if "ok" in s or "complet" in s or "termin" in s or "done" in s or "clos" in s: item["status"] = "Completed"
            elif "progress" in s or "cours" in s or "ongoing" in s: item["status"] = "In Progress"
            elif "risk" in s or "risque" in s or "delay" in s: item["status"] = "At Risk"
            elif "block" in s or "bloqu" in s or "stop" in s: item["status"] = "Blocked"
            elif "track" in s: item["status"] = "On Track"
            elif "cancel" in s or "annul" in s: item["status"] = "Cancelled"
        elif col_progress and not pd.isna(row.get(col_progress)):
            try:
                prog_val = float(str(row.get(col_progress)).replace('%', '').strip())
                if prog_val > 1.0: prog_val = prog_val / 100.0
                if prog_val >= 1.0:
                    item["status"] = "Completed"
                elif prog_val > 0.0:
                    item["status"] = "In Progress"
                else:
                    item["status"] = "Not Started"
            except (ValueError, TypeError):
                pass
            
        if col_priority and not pd.isna(row.get(col_priority)):
            p = str(row.get(col_priority)).strip().lower()
            if "high" in p or "haute" in p or "crit" in p or "1" in p: item["priority"] = "High"
            elif "low" in p or "basse" in p or "3" in p: item["priority"] = "Low"
            
        items.append(item)
        
    if not items:
        raise ValueError("Aucun item valide trouvé dans ce fichier. Vérifiez les noms des colonnes.")
        
    return items
