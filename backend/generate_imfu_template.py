import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

def create_sabah_based_template(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    
    # 1. Keep only the main sheet
    main_sheet_name = wb.sheetnames[0]
    for sn in wb.sheetnames:
        if sn != main_sheet_name:
            del wb[sn]
            
    ws = wb.active
    ws.title = "IMFU"
    
    # 2. Clear data from row 7 downwards (Sabah's data starts at 7)
    max_r = ws.max_row
    if max_r >= 7:
        ws.delete_rows(7, max_r - 6)
        
    # 3. Clean "Unnamed" headers in rows 5 and 6
    max_c = ws.max_column
    for r in [5, 6]:
        for c in range(1, max_c + 1):
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str) and 'Unnamed' in val:
                ws.cell(row=r, column=c, value='')
                
    # 4. Append Badr's columns at the end
    badr_cols = [
        "Verif Modifs & GAP", "Validation Livrables", "Affectation Matière",
        "Gamme Fabrication", "Crosscheck TTBOM", "ELSA Injections",
        "MAJ Jigboard (Badr)", "Premier Boût", "Sertissage centralisé",
        "MAJ Gamme Assemblage", "Ajustement longueurs", "Train Me",
        "Optimisation longueurs", "Coupe gaine", "Task Seq Cheminement",
        "Impression Gamme", "Go PME Cheminement", "Task Seq Intégration",
        "Prep FI", "Go PME Intégration"
    ]
    
    start_c = max_c + 1
    # Add a main group header for Badr at row 5 (Sabah's group row)
    ws.merge_cells(start_row=5, start_column=start_c, end_row=5, end_column=start_c + len(badr_cols) - 1)
    grp_cell = ws.cell(row=5, column=start_c, value="MANUFACTURING (BADR)")
    grp_cell.font = Font(name='Calibri', bold=True, size=11, color="000000")
    grp_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type='solid')
    grp_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(left=Side(style='thin', color='A0A0A0'), right=Side(style='thin', color='A0A0A0'), top=Side(style='thin', color='A0A0A0'), bottom=Side(style='thin', color='A0A0A0'))
    
    for i, col_name in enumerate(badr_cols):
        c_idx = start_c + i
        ws.cell(row=5, column=c_idx).border = thin_border
        
        cell = ws.cell(row=6, column=c_idx, value=col_name)
        cell.font = Font(name='Calibri', bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="001F3F", end_color="001F3F", fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(c_idx)].width = 20
        
    # 5. Apply clean formatting and borders to first 100 rows
    total_cols = start_c + len(badr_cols) - 1
    for r in range(7, 107):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            
    # Enable filter
    ws.auto_filter.ref = f"A6:{get_column_letter(total_cols)}107"
    
    # 6. Quality of Life Improvements
    # Freeze panes: Lock rows 1-6 (headers) and columns A-G (Sequence to Name)
    ws.freeze_panes = "H7"
    
    # Hide default gridlines for a cleaner look
    ws.sheet_view.showGridLines = False

    # Add Data Validation (Drop-downs)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Blocked,At Risk,Completed"', allow_blank=True)
    
    # Find columns to apply validation/formatting
    status_cols = []
    progress_cols = []
    for c in range(1, total_cols + 1):
        val = str(ws.cell(row=6, column=c).value).lower()
        if 'status' in val:
            status_cols.append(c)
        elif 'progress' in val:
            progress_cols.append(c)

    for c in status_cols:
        col_letter = get_column_letter(c)
        dv_status.add(f'{col_letter}7:{col_letter}200')
    ws.add_data_validation(dv_status)

    # Add Conditional Formatting
    from openpyxl.formatting.rule import CellIsRule, DataBarRule
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    for c in status_cols:
        col_letter = get_column_letter(c)
        ws.conditional_formatting.add(f'{col_letter}7:{col_letter}200', CellIsRule(operator='equal', formula=['"Completed"'], fill=green_fill))
        ws.conditional_formatting.add(f'{col_letter}7:{col_letter}200', CellIsRule(operator='equal', formula=['"Blocked"'], fill=red_fill))
        ws.conditional_formatting.add(f'{col_letter}7:{col_letter}200', CellIsRule(operator='equal', formula=['"In Progress"'], fill=yellow_fill))

    for c in progress_cols:
        col_letter = get_column_letter(c)
        # Add a nice data bar to progress columns
        dbar = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color="63C384", showValue="None", minLength=None, maxLength=None)
        ws.conditional_formatting.add(f'{col_letter}7:{col_letter}200', dbar)
    
    wb.save(output_path)
    print(f"✅ Unified Template saved to {output_path} with {total_cols} columns")

if __name__ == "__main__":
    create_sabah_based_template(
        "/Users/yasser/Desktop/Dasboard ALSTOM/IMFU_Sabah.xlsx",
        "/Users/yasser/Desktop/Dasboard ALSTOM/IMFU_NEW_STANDARD.xlsx"
    )
